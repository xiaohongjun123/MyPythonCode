from openpyxl import load_workbook
class ParaExcel(object):
    def __init__(self,excelpath,sheetName):
        self.wb=load_workbook(excelpath)
        self.sheet=self.wb.get_sheet_by_name(sheetName)
        self.maxRowNum=self.sheet.max_row

    def getDataFromSheet(self):
        datalist=[]
        for line in self.sheet.rows[1:]:
            temlist=[]
            temlist.append(line[0].value)
            temlist.append(line[1].value)
            temlist.append(line[2].value)
            datalist.append(temlist)
        return datalist

if __name__=="__main__":
    Data=ParaExcel(r"G:\Test.xlsx","Sheet1")
    print(Data.getDataFromSheet())


