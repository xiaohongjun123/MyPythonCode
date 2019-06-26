from openpyxl import load_workbook
class ParaExcel(object):
    def __init__(self,excelpath,sheetname):
        self.wb=load_workbook(excelpath)
        self.sheet=self.wb.get_sheet_by_name(sheetname)
        self.maxRowNum=self.sheet.max_row

    def getDataFromSheet(self):
        datalist=[]
        for line in self.sheet.rows[1:]:
            list=[]
            list.append(line[1].value)
            list.append(line[2].value)
            list.append(line[3].value)
            datalist.append(list)
        return datalist

if __name__=="__main__":
    a=ParaExcel(r"G:\addemployee.xlsx","Sheet1")
    print(a.getDataFromSheet())
